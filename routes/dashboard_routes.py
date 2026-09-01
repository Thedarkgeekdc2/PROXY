import io
import calendar
from collections import defaultdict
from datetime import datetime, date, timedelta
from flask import Blueprint, render_template, request, send_file, jsonify
from services.report_service import (
    get_dashboard_summary, get_daily_report, get_custom_report,
    get_weekly_report, get_monthly_report, get_teacher_load_report, get_merge_report
)
from database.models import db, Teacher, ProxyAssignment, Absence

dashboard_bp = Blueprint('dashboard', __name__)


@dashboard_bp.route('/api/teachers')
def api_teachers():
    teachers = Teacher.query.filter_by(status='active').order_by(Teacher.name).all()
    return jsonify([{
        'id': t.id,
        'name': t.name,
        'full_name': t.full_name or '',
        'subject': t.subject or '',
    } for t in teachers])


@dashboard_bp.route('/api/teachers/<int:teacher_id>/update', methods=['POST'])
def api_update_teacher(teacher_id):
    teacher = Teacher.query.get_or_404(teacher_id)
    data = request.get_json(force=True)
    if 'full_name' in data:
        teacher.full_name = data['full_name'].strip() or None
    db.session.commit()
    return jsonify({'success': True, 'full_name': teacher.full_name or '', 'display_name': teacher.display_name})


@dashboard_bp.route('/api/class_timetable')
def api_class_timetable():
    """P1-P8 teachers for a class on a given date, with proxy substitutions shown."""
    from database.models import Class, Timetable
    class_id = request.args.get('class_id', type=int)
    date_str = request.args.get('date', date.today().strftime('%Y-%m-%d'))
    if not class_id:
        return jsonify({'error': 'No class selected', 'periods': []})

    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        target_date = date.today()

    day = target_date.strftime('%A')
    cls = Class.query.get(class_id)
    if not cls:
        return jsonify({'error': 'Class not found', 'periods': []})

    tts = {tt.period_no: tt for tt in
           Timetable.query.filter_by(class_id=class_id, day=day).all()}

    proxies = {p.period_no: p for p in
               ProxyAssignment.query.filter_by(
                   class_id=class_id, date=target_date, status='confirmed').all()}

    periods = []
    for p in range(1, 9):
        tt    = tts.get(p)
        proxy = proxies.get(p)
        if proxy and proxy.proxy_teacher:
            teacher_name = proxy.proxy_teacher.display_name
            is_proxy     = True
        elif tt and tt.teacher:
            teacher_name = tt.teacher.display_name
            is_proxy     = False
        else:
            teacher_name = None
            is_proxy     = False
        periods.append({
            'period':  p,
            'teacher': teacher_name,
            'is_proxy': is_proxy,
        })

    return jsonify({'class_label': cls.label, 'day': day, 'periods': periods})


@dashboard_bp.route('/')
def index():
    from database.models import Class
    summary     = get_dashboard_summary()
    all_classes = Class.query.order_by(Class.class_name, Class.section).all()
    return render_template('dashboard.html', summary=summary, all_classes=all_classes)


@dashboard_bp.route('/reports')
def reports():
    tab = request.args.get('tab', 'daily')
    date_str = request.args.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        ref_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        ref_date = date.today()

    daily        = get_daily_report(ref_date)
    weekly       = get_weekly_report(ref_date)
    monthly      = get_monthly_report(ref_date)
    teacher_load = get_teacher_load_report()
    merges       = get_merge_report()

    # Custom date range
    from_str = request.args.get('from_date', date_str)
    to_str   = request.args.get('to_date',   date_str)
    try:
        from_date = datetime.strptime(from_str, '%Y-%m-%d').date()
    except Exception:
        from_date = ref_date
    try:
        to_date = datetime.strptime(to_str, '%Y-%m-%d').date()
    except Exception:
        to_date = ref_date

    custom = get_custom_report(from_date, to_date)

    return render_template('reports.html',
                           tab=tab, ref_date=ref_date, date_str=date_str,
                           from_str=from_str, to_str=to_str,
                           daily=daily, custom=custom, weekly=weekly, monthly=monthly,
                           teacher_load=teacher_load, merges=merges)


@dashboard_bp.route('/export/excel')
def export_excel():
    """Export proxy report as formatted .xlsx"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        return "openpyxl not installed", 500

    tab   = request.args.get('tab', 'daily')
    today = date.today()

    # Shared styles
    HDR_FILL  = PatternFill('solid', fgColor='1E3A5F')
    HDR_FONT  = Font(bold=True, color='FFFFFF', size=10)
    TTL_FONT  = Font(bold=True, size=13, color='1E3A5F')
    SUB_FONT  = Font(italic=True, color='64748B', size=9)
    ALT_FILL  = PatternFill('solid', fgColor='F0F4FA')
    PRX_FILL  = PatternFill('solid', fgColor='D1FAE5')
    MRG_FILL  = PatternFill('solid', fgColor='EDE9FE')
    ABS_FILL  = PatternFill('solid', fgColor='FEE2E2')
    thin = Border(
        left=Side(style='thin',   color='CBD5E1'),
        right=Side(style='thin',  color='CBD5E1'),
        top=Side(style='thin',    color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left', vertical='center', wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Proxy Report'

    def set_title(title_text, ncols):
        from openpyxl.utils import get_column_letter
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(1, 1, title_text); c.font = TTL_FONT; c.alignment = center
        ws.row_dimensions[1].height = 26
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        c = ws.cell(2, 1, f'Generated: {today.strftime("%d %B %Y")}   |   KVS Proxy System')
        c.font = SUB_FONT; c.alignment = center
        ws.row_dimensions[2].height = 14

    def style_hdr_row(hdr_row, ncols):
        ws.row_dimensions[hdr_row].height = 20
        for col in range(1, ncols + 1):
            cell = ws.cell(hdr_row, col)
            cell.fill = HDR_FILL; cell.font = HDR_FONT
            cell.alignment = center; cell.border = thin

    def style_data(row, ncols, fill=None, left_cols=None):
        for col in range(1, ncols + 1):
            cell = ws.cell(row, col)
            if fill: cell.fill = fill
            cell.border = thin
            cell.alignment = left if (left_cols and col in left_cols) else center
        ws.row_dimensions[row].height = 16

    # ── DAILY ──────────────────────────────────────────────────────────────
    if tab == 'daily':
        date_str = request.args.get('date', today.strftime('%Y-%m-%d'))
        try:
            ref_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except Exception:
            ref_date = today
        absences = Absence.query.filter_by(absent_date=ref_date).all()
        proxies  = ProxyAssignment.query.filter_by(date=ref_date).order_by(
                       ProxyAssignment.period_no).all()
        title = f'Daily Proxy Report — {ref_date.strftime("%d %B %Y")}'
        fname = f'proxy_daily_{ref_date}.xlsx'
        ncols = 8
        set_title(title, ncols)
        # Absent teachers row
        ws.append([])  # row 3
        ws.cell(4, 1, 'ABSENT TEACHERS').font = Font(bold=True, color='DC2626', size=10)
        ws.row_dimensions[4].height = 18
        r5 = 5
        ws.merge_cells(start_row=r5, start_column=1, end_row=r5, end_column=ncols)
        if absences:
            names = ',  '.join(ab.teacher.display_name for ab in absences)
            c = ws.cell(r5, 1, names)
            c.fill = ABS_FILL; c.font = Font(bold=True, color='991B1B')
            c.alignment = left; c.border = thin
        else:
            ws.cell(r5, 1, 'No absences on this date').font = Font(italic=True, color='64748B')
        ws.row_dimensions[r5].height = 18
        # Header
        hdr_r = 7
        for ci, h in enumerate(['Period','Class','Absent Teacher','Proxy Teacher','Status','Score','Reason','Day'], 1):
            ws.cell(hdr_r, ci, h)
        style_hdr_row(hdr_r, ncols)
        # Data
        pc = mc = 0
        for i, p in enumerate(proxies):
            r = hdr_r + 1 + i
            ws.cell(r, 1, f'P{p.period_no}')
            ws.cell(r, 2, p.class_.label if p.class_ else '—')
            ws.cell(r, 3, p.original_teacher.display_name if p.original_teacher else '—')
            prx = p.proxy_teacher.display_name if p.proxy_teacher else ('CLASS MERGE' if p.status=='merge' else '—')
            ws.cell(r, 4, prx)
            ws.cell(r, 5, (p.status or '').upper())
            ws.cell(r, 6, p.score or 0)
            ws.cell(r, 7, p.reason or '—')
            ws.cell(r, 8, p.day or '—')
            fill = PRX_FILL if p.status=='confirmed' else (MRG_FILL if p.status=='merge' else (ALT_FILL if i%2 else None))
            style_data(r, ncols, fill=fill, left_cols={3,4,7})
            if p.status=='confirmed': pc += 1
            # merge count comes from ClassDayMerge, not proxy status
        if not proxies:
            r = hdr_r + 1
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            ws.cell(r, 1, 'No proxy records for this date').font = Font(italic=True, color='64748B')
        # Summary
        sr = hdr_r + len(proxies) + 3
        ws.cell(sr, 1, f'Absent: {len(absences)}').font = Font(bold=True, color='DC2626')
        ws.cell(sr, 3, f'Confirmed: {pc}').font = Font(bold=True, color='059669')
        ws.cell(sr, 5, f'Merges: {mc}').font = Font(bold=True, color='7C3AED')
        ws.cell(sr, 7, f'Pending: {len(proxies)-pc-mc}').font = Font(bold=True, color='D97706')
        for ci, w in zip(range(1,9), [8,9,24,24,12,8,40,12]):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(ci)].width = w

    # ── CUSTOM DATE RANGE ──────────────────────────────────────────────────
    elif tab == 'custom':
        from_str = request.args.get('from_date', today.strftime('%Y-%m-%d'))
        to_str   = request.args.get('to_date',   today.strftime('%Y-%m-%d'))
        try:
            from_date = datetime.strptime(from_str, '%Y-%m-%d').date()
        except Exception:
            from_date = today
        try:
            to_date = datetime.strptime(to_str, '%Y-%m-%d').date()
        except Exception:
            to_date = today
        if to_date < from_date:
            to_date = from_date
        delta = (to_date - from_date).days + 1
        dates = [from_date + timedelta(days=i) for i in range(delta)]
        all_teachers = Teacher.query.filter_by(status='active').order_by(Teacher.name).all()
        proxies = ProxyAssignment.query.filter(
            ProxyAssignment.date >= from_date,
            ProxyAssignment.date <= to_date,
            ProxyAssignment.status == 'confirmed'
        ).all()
        counts = defaultdict(lambda: defaultdict(int))
        for p in proxies:
            if p.proxy_teacher_id:
                counts[p.proxy_teacher_id][p.date] += 1
        title = f'Custom Report — {from_date.strftime("%d/%m/%Y")} to {to_date.strftime("%d/%m/%Y")}'
        fname = f'proxy_custom_{from_date}_to_{to_date}.xlsx'
        ncols = 2 + len(dates) + 1
        set_title(title, ncols)
        hdr_r = 4
        ws.cell(hdr_r, 1, '#')
        ws.cell(hdr_r, 2, 'Teacher')
        for di, d in enumerate(dates):
            ws.cell(hdr_r, 3+di, d.strftime('%d/%m'))
        ws.cell(hdr_r, 3+len(dates), 'TOTAL')
        style_hdr_row(hdr_r, ncols)
        grand = 0
        for ri, t in enumerate(all_teachers):
            r = hdr_r + 1 + ri
            ws.cell(r, 1, ri+1)
            ws.cell(r, 2, t.display_name)
            row_total = 0
            for di, d in enumerate(dates):
                cnt = counts[t.id].get(d, 0)
                c = ws.cell(r, 3+di, cnt if cnt else '')
                if cnt > 0: c.font = Font(bold=True, color='059669')
                row_total += cnt
            tc = ws.cell(r, 3+len(dates), row_total if row_total else '')
            if row_total: tc.font = Font(bold=True, color='1E3A5F')
            grand += row_total
            style_data(r, ncols, fill=(ALT_FILL if ri%2 else None), left_cols={2})
        gr = hdr_r + len(all_teachers) + 2
        ws.cell(gr, 2, 'GRAND TOTAL').font = Font(bold=True, color='1E3A5F')
        ws.cell(gr, 3+len(dates), grand).font = Font(bold=True, color='059669', size=11)
        from openpyxl.utils import get_column_letter
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 26
        for di in range(len(dates)):
            ws.column_dimensions[get_column_letter(3+di)].width = 9
        ws.column_dimensions[get_column_letter(3+len(dates))].width = 10

    # ── MONTHLY ────────────────────────────────────────────────────────────
    else:
        date_str = request.args.get('date', today.strftime('%Y-%m-%d'))
        try:
            ref_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except Exception:
            ref_date = today
        year  = ref_date.year
        month = ref_date.month
        _, ld = calendar.monthrange(year, month)
        m_start = date(year, month, 1)
        m_end   = date(year, month, ld)   # full month
        dates   = [m_start + timedelta(days=i) for i in range(ld)]
        all_teachers = Teacher.query.filter_by(status='active').order_by(Teacher.name).all()
        proxies = ProxyAssignment.query.filter(
            ProxyAssignment.date >= m_start,
            ProxyAssignment.date <= m_end,
            ProxyAssignment.status == 'confirmed'
        ).all()
        counts = defaultdict(lambda: defaultdict(int))
        for p in proxies:
            if p.proxy_teacher_id:
                counts[p.proxy_teacher_id][p.date] += 1

        # Leave Summary: for each teacher, how many periods needed a confirmed
        # proxy this month because THEY were absent — i.e. count of confirmed
        # ProxyAssignment rows where they are the ORIGINAL (absent) teacher.
        # This updates live as proxies get confirmed (grows only once a
        # period's proxy is actually assigned, never before).
        leave_source = ProxyAssignment.query.filter(
            ProxyAssignment.date >= m_start,
            ProxyAssignment.date <= m_end,
            ProxyAssignment.status == 'confirmed'
        ).all()
        leave_counts = defaultdict(int)
        for p in leave_source:
            if p.original_teacher_id:
                leave_counts[p.original_teacher_id] += 1

        title = f'Monthly Proxy Report — {calendar.month_name[month]} {year}'
        fname = f'proxy_monthly_{year}_{month:02d}.xlsx'
        ncols = 2 + len(dates) + 1 + 1   # +1 TOTAL, +1 LEAVE SUMMARY
        set_title(title, ncols)
        hdr_r = 4
        ws.cell(hdr_r, 1, '#')
        ws.cell(hdr_r, 2, 'Teacher')
        for di, d in enumerate(dates):
            ws.cell(hdr_r, 3+di, d.strftime('%d'))
        ws.cell(hdr_r, 3+len(dates), 'TOTAL')
        ws.cell(hdr_r, 3+len(dates)+1, 'LEAVE SUMMARY')
        style_hdr_row(hdr_r, ncols)
        grand       = 0
        grand_leave = 0
        for ri, t in enumerate(all_teachers):
            r = hdr_r + 1 + ri
            ws.cell(r, 1, ri+1)
            ws.cell(r, 2, t.display_name)
            row_total = 0
            for di, d in enumerate(dates):
                cnt = counts[t.id].get(d, 0)
                c = ws.cell(r, 3+di, cnt if cnt else '')
                if cnt > 0: c.font = Font(bold=True, color='059669')
                row_total += cnt
            tc = ws.cell(r, 3+len(dates), row_total if row_total else '')
            if row_total: tc.font = Font(bold=True, color='1E3A5F')
            grand += row_total

            lv = leave_counts.get(t.id, 0)
            lc = ws.cell(r, 3+len(dates)+1, lv if lv else '')
            if lv: lc.font = Font(bold=True, color='B91C1C')
            grand_leave += lv

            style_data(r, ncols, fill=(ALT_FILL if ri%2 else None), left_cols={2})
        gr = hdr_r + len(all_teachers) + 2
        ws.cell(gr, 2, 'GRAND TOTAL').font = Font(bold=True, color='1E3A5F')
        ws.cell(gr, 3+len(dates), grand).font = Font(bold=True, color='059669', size=11)
        ws.cell(gr, 3+len(dates)+1, grand_leave).font = Font(bold=True, color='B91C1C', size=11)
        from openpyxl.utils import get_column_letter
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 26
        for di in range(len(dates)):
            ws.column_dimensions[get_column_letter(3+di)].width = 5
        ws.column_dimensions[get_column_letter(3+len(dates))].width = 10
        ws.column_dimensions[get_column_letter(3+len(dates)+1)].width = 15

    # Real merge count from ClassDayMerge
    from database.models import ClassDayMerge as _CDM
    if tab == 'daily':
        try:
            from datetime import datetime as _dt
            _ref = _dt.strptime(request.args.get('date', today.strftime('%Y-%m-%d')), '%Y-%m-%d').date()
        except Exception:
            _ref = today
        mc = _CDM.query.filter_by(date=_ref).count()
    elif tab == 'weekly':
        mc = _CDM.query.filter(
            _CDM.date >= week_start, _CDM.date <= week_end).count() if 'week_start' in dir() else 0
    elif tab == 'monthly':
        mc = _CDM.query.filter(
            _CDM.date >= m_start, _CDM.date <= m_end).count() if 'm_start' in dir() else 0

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=fname
    )

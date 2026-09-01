"""
Excel parser for KVS Master Timetable format (Master_TT.xlsx).

Master_TT layout (0-indexed columns):
  Col 0      : Class label  (IA, IB, IIA … VD)
  Col 97     : Class label  (repeated – ignored)
  Row 0      : Day headers  MON@1, TUE@17, WED@33, THU@49, FRI@65, SAT@81
  Row 1      : Period nos   1-8 repeating per day
  Row 2      : TR / SUB labels
  Rows 3-22  : Data rows (one per class)
  Row 23     : Footer number sequence – skipped

For each day the 8 periods occupy 16 consecutive columns:
  Period p (1-8) → col = day_start + (p-1)*2       → TR (teacher code)
                   col = day_start + (p-1)*2 + 1    → SUB (subject)
"""

from openpyxl import load_workbook
from database.models import db, Teacher, Class, Timetable

# ── Constants ──────────────────────────────────────────────────────────────
DAY_MAP = {
    'MON': 'Monday', 'TUE': 'Tuesday', 'WED': 'Wednesday',
    'THU': 'Thursday', 'FRI': 'Friday', 'SAT': 'Saturday'
}

# Day abbreviation → 0-indexed starting column in Master_TT
MASTER_DAY_STARTS = {
    'MON': 1,
    'TUE': 17,
    'WED': 33,
    'THU': 49,
    'FRI': 65,
    'SAT': 81,
}

# Rows that are headers or footers (0-indexed)
MASTER_SKIP_ROWS = {0, 1, 2, 23}

VALID_CLASS_PREFIXES = ('I', 'V')   # quick sanity check


# ── Helpers ────────────────────────────────────────────────────────────────
def _roman_to_num(s):
    table = [('III', '3'), ('IV', '4'), ('VII', '7'), ('VIII', '8'),
             ('VI', '6'), ('II', '2'), ('V', '5'), ('I', '1')]
    for roman, num in table:
        if s.startswith(roman):
            return num, s[len(roman):]
    return s, ''


def _parse_class_label(label):
    """'IVA' → ('4', 'A'), 'IIIB' → ('3', 'B')"""
    label = str(label).strip().upper()
    num, rest = _roman_to_num(label)
    return num, rest.strip() if rest else ''


def _v(val, default=''):
    """Clean a raw cell value."""
    if val is None:
        return default
    s = str(val).strip()
    return default if s in ('', '0', '0.0', 'None', 'nan') else s


def _safe_rows(ws):
    return list(ws.iter_rows(values_only=True))


def _is_valid_class(label):
    """Returns True if label looks like a real class (IA, IIB, VA ...)."""
    if not label:
        return False
    s = str(label).strip().upper()
    return any(s.startswith(p) for p in VALID_CLASS_PREFIXES) and len(s) <= 5


# ── Master TT parser ───────────────────────────────────────────────────────
def parse_mastertt(file_path):
    """
    Parse Master_TT.xlsx.

    Returns list of records:
        [{'class_label', 'day', 'period_no', 'teacher_code', 'subject'}]
    """
    wb   = load_workbook(file_path, read_only=True)
    ws   = wb.active
    rows = _safe_rows(ws)

    records = []

    for row_idx, row in enumerate(rows):
        if row_idx in MASTER_SKIP_ROWS:
            continue

        class_label = _v(row[0] if row else None)
        if not _is_valid_class(class_label):
            continue

        for day_abbr, day_start in MASTER_DAY_STARTS.items():
            full_day = DAY_MAP[day_abbr]
            for period_no in range(1, 9):
                tr_col  = day_start + (period_no - 1) * 2
                sub_col = tr_col + 1

                teacher_code = _v(row[tr_col]  if len(row) > tr_col  else None)
                subject      = _v(row[sub_col] if len(row) > sub_col else None)

                if teacher_code:
                    records.append({
                        'class_label':  class_label,
                        'day':          full_day,
                        'period_no':    period_no,
                        'teacher_code': teacher_code,
                        'subject':      subject,
                    })

    return records


# ── DB helpers ─────────────────────────────────────────────────────────────
def _upsert_teacher(code, subject=''):
    t = Teacher.query.filter_by(name=code).first()
    if not t:
        t = Teacher(name=code, subject=subject, level='both', max_daily_proxy=2)
        db.session.add(t)
        db.session.flush()
    return t


def _upsert_class(class_label):
    num, section = _parse_class_label(class_label)
    if not num or not section:
        return None
    cls = Class.query.filter_by(class_name=num, section=section).first()
    if not cls:
        cls = Class(class_name=num, section=section)
        db.session.add(cls)
        db.session.flush()
    return cls


# ── Main upload function ───────────────────────────────────────────────────
def upload_mastertt_excel(file_path):
    """
    Parse and save Master_TT.xlsx into the database.

    Returns:
        (teacher_count, tt_count, errors)
    """
    try:
        records = parse_mastertt(file_path)
    except Exception as e:
        return 0, 0, [f'File read error: {str(e)}']

    teacher_cache = {}
    class_cache   = {}
    tt_count      = 0
    errors        = []

    for rec in records:
        code    = rec['teacher_code']
        subject = rec['subject']

        # Upsert teacher
        if code not in teacher_cache:
            t = Teacher.query.filter_by(name=code).first()
            if not t:
                t = Teacher(name=code, subject=subject, level='both', max_daily_proxy=2)
                db.session.add(t)
                db.session.flush()
            teacher_cache[code] = t
        teacher = teacher_cache[code]

        # Upsert class
        lbl = rec['class_label']
        if lbl not in class_cache:
            cls = _upsert_class(lbl)
            class_cache[lbl] = cls
        cls = class_cache[lbl]

        if not cls:
            errors.append(f'Cannot parse class: {lbl}')
            continue

        # Skip duplicates
        existing = Timetable.query.filter_by(
            day=rec['day'], period_no=rec['period_no'], class_id=cls.id
        ).first()
        if existing:
            continue

        tt = Timetable(
            day=rec['day'],
            period_no=rec['period_no'],
            class_id=cls.id,
            teacher_id=teacher.id,
            subject=subject,
        )
        db.session.add(tt)
        tt_count += 1

    db.session.commit()
    teacher_count = len(teacher_cache)
    return teacher_count, tt_count, errors

